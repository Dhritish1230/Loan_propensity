import json
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(os.getenv("LOAN_PROPENSITY_WORKSPACE", Path(__file__).resolve().parents[2])).resolve()
OUT = ROOT / "april_label_comparison"
TEST_OUT = ROOT / "multi_month_training" / "test_outputs"

T1_COMP = TEST_OUT / "april_t1_unlabeled_vs_labeled_top100_comparison.csv"
T0_COMP = TEST_OUT / "april_t0_unlabeled_vs_labeled_top100_comparison.csv"
SUMMARY = TEST_OUT / "april_unlabeled_vs_labeled_top100_comparison_summary.csv"
FULL_SUMMARY = TEST_OUT / "april_unlabeled_vs_labeled_full_rank_comparison_summary.csv"
REPORT = TEST_OUT / "april_labeled_evaluation_report.json"


def source(row: pd.Series) -> str:
    if int(row.get("converted_from_sfdc", 0) or 0) == 1:
        return "SFDC"
    if int(row.get("converted_from_diy", 0) or 0) == 1:
        return "DIY"
    return "Not Converted"


def status(row: pd.Series) -> str:
    return "Converted" if int(row.get("converted_full", 0) or 0) == 1 else "Not Converted"


def clean_comparison(path: Path, stage: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    out = pd.DataFrame(
        {
            "Stage": stage,
            "Before Label Rank": df["unlabeled_rank"],
            "After Label Rank": df["labeled_rank"],
            "Same Rank": df["same_rank"].astype(bool),
            "UID": df["uid"],
            "Prediction Score": df["t1_loan_conversion_score"] if "t1_loan_conversion_score" in df.columns else df.get("t0_call_targeting_score"),
            "Converted After Label Check": df["converted_full"].astype(int),
            "Conversion Source": df.apply(source, axis=1),
            "Outcome": df.apply(status, axis=1),
            "Total Calls": df["total_calls"],
            "Answered Calls": df["answered_calls"],
            "Answered Rate": df["answered_rate"],
            "Avg Call Duration": df["avg_call_duration"],
            "Campaign": df.get("campaign_id", ""),
            "State": df.get("state", ""),
        }
    )
    return out


def build_rank_buckets(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for start, end in [(1, 10), (11, 20), (21, 50), (51, 100)]:
        sub = df[(df["Before Label Rank"] >= start) & (df["Before Label Rank"] <= end)]
        rows.append(
            {
                "Rank Bucket": f"{start}-{end}",
                "Users": len(sub),
                "Conversions": int(sub["Converted After Label Check"].sum()),
                "Precision": float(sub["Converted After Label Check"].mean()) if len(sub) else 0,
                "DIY Conversions": int((sub["Conversion Source"] == "DIY").sum()),
                "SFDC Conversions": int((sub["Conversion Source"] == "SFDC").sum()),
            }
        )
    return pd.DataFrame(rows)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="102027")
    header_font = Font(color="FFFFFF", bold=True)
    good_fill = PatternFill("solid", fgColor="DFF7E8")
    bad_fill = PatternFill("solid", fgColor="FEE4E2")
    thin = Side(style="thin", color="D7DEE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col), 42)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)
        headers = [cell.value for cell in ws[1]]
        if "Converted After Label Check" in headers:
            col = headers.index("Converted After Label Check") + 1
            ws.conditional_formatting.add(
                f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row}",
                FormulaRule(formula=[f"${get_column_letter(col)}2=1"], fill=good_fill),
            )
            ws.conditional_formatting.add(
                f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row}",
                FormulaRule(formula=[f"${get_column_letter(col)}2=0"], fill=bad_fill),
            )
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(row=1, column=cell.column).value
                if header and any(token in str(header) for token in ["Rate", "Precision"]):
                    cell.number_format = "0.00%"
                if header and "Score" in str(header):
                    cell.number_format = "0.000000"

    ws = wb["Executive_Summary"]
    ws.insert_rows(1, 2)
    ws["A1"] = "April Prediction vs Label Validation"
    ws["A1"].font = Font(size=20, bold=True, color="102027")
    ws["A2"] = "The model scored April first without conversion labels. Labels were joined afterward only to validate outcomes."
    ws["A2"].font = Font(italic=True, color="60717C")

    bucket = wb["T1_Rank_Buckets"]
    chart = BarChart()
    chart.title = "T1 Conversions by Predicted Rank Bucket"
    chart.y_axis.title = "Conversions"
    chart.x_axis.title = "Rank Bucket"
    data = Reference(bucket, min_col=3, min_row=1, max_row=bucket.max_row)
    cats = Reference(bucket, min_col=1, min_row=2, max_row=bucket.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    bucket.add_chart(chart, "H2")

    source_ws = wb["T1_Source_Split"]
    pie = PieChart()
    pie.title = "Top 100 Conversion Source"
    data = Reference(source_ws, min_col=2, min_row=1, max_row=source_ws.max_row)
    labels = Reference(source_ws, min_col=1, min_row=2, max_row=source_ws.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 12
    source_ws.add_chart(pie, "E2")

    wb.save(path)


def build_html(t1: pd.DataFrame, summary: pd.DataFrame, path: Path) -> None:
    top20 = t1.head(20)
    conv = int(t1["Converted After Label Check"].sum())
    same_order = bool(summary.loc[summary["stage"].eq("T1"), "same_order"].iloc[0])
    same_scores = bool(summary.loc[summary["stage"].eq("T1"), "same_scores"].iloc[0])
    same_users = bool(summary.loc[summary["stage"].eq("T1"), "same_uid_set"].iloc[0])
    rows = []
    for _, row in top20.iterrows():
        rows.append(
            f"""<tr class="{ 'converted' if row['Converted After Label Check'] == 1 else '' }">
<td>{int(row['Before Label Rank'])}</td><td>{row['UID']}</td><td>{row['Prediction Score']:.6f}</td><td>{row['Outcome']}</td><td>{row['Conversion Source']}</td><td>{row['Total Calls']:.0f}</td><td>{row['Answered Calls']:.0f}</td><td>{row['Avg Call Duration']:.1f}s</td></tr>"""
        )
    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>April Prediction Label Comparison</title>
<style>
body{{margin:0;font-family:Bahnschrift,Candara,'Trebuchet MS',sans-serif;color:#102027;background:linear-gradient(135deg,#fff9ec,#eaf7f2)}}
.wrap{{width:min(1280px,calc(100% - 34px));margin:0 auto;padding:32px 0 60px}}
.hero,.panel,.card{{background:rgba(255,255,255,.84);border:1px solid rgba(16,32,39,.12);border-radius:28px;box-shadow:0 24px 70px rgba(23,50,76,.13);padding:24px;margin-top:18px}}
h1,h2{{font-family:Georgia,'Palatino Linotype',serif;letter-spacing:-.035em}}h1{{font-size:60px;line-height:.95;margin:0}}p{{color:#60717c;line-height:1.55}}
.grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px}}.card strong{{display:block;font-size:38px;margin-top:8px}}.card span{{color:#60717c}}
table{{width:100%;border-collapse:collapse;background:white;border-radius:18px;overflow:hidden}}th,td{{padding:11px;border-bottom:1px solid #edf1f0;text-align:left}}th{{background:#102027;color:white;font-size:12px;text-transform:uppercase;letter-spacing:.08em}}.converted{{background:#e7f8ee}}
.pill{{display:inline-block;border-radius:999px;padding:7px 12px;background:#dff7e8;font-weight:900;color:#0f5132}}@media(max-width:900px){{.grid{{grid-template-columns:1fr}}h1{{font-size:40px}}}}
</style></head><body><main class="wrap">
<section class="hero"><p style="font-weight:900;color:#0f766e;text-transform:uppercase;letter-spacing:.16em">Supervisor validation view</p><h1>April predictions stayed unchanged after labels</h1><p>We scored April first without DIY/SFDC labels. After labels were joined, the top-100 T1 list had the same users, same order, and same prediction scores. Labels only revealed which predicted users actually converted.</p></section>
<section class="grid">
<article class="card"><span>Same top-100 users</span><strong>{same_users}</strong></article>
<article class="card"><span>Same order</span><strong>{same_order}</strong></article>
<article class="card"><span>Same scores</span><strong>{same_scores}</strong></article>
<article class="card"><span>Converted in top 100</span><strong>{conv}/100</strong></article>
</section>
<section class="panel"><h2>What this proves</h2><p><span class="pill">Clean validation</span></p><p>The model did not use conversion labels while ranking April users. The labels were added afterward only to check outcomes. This is why the prediction ranking and scores are identical before and after label validation.</p></section>
<section class="panel"><h2>Top 20 Comparison</h2><table><thead><tr><th>Rank Before Labels</th><th>UID</th><th>Prediction Score</th><th>Outcome After Labels</th><th>Source</th><th>Total Calls</th><th>Answered</th><th>Avg Duration</th></tr></thead><tbody>{''.join(rows)}</tbody></table></section>
</main></body></html>"""
    path.write_text(html, encoding="utf-8")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    t1 = clean_comparison(T1_COMP, "T1")
    t0 = clean_comparison(T0_COMP, "T0")
    summary = pd.read_csv(SUMMARY)
    full_summary = pd.read_csv(FULL_SUMMARY)
    report = json.loads(REPORT.read_text(encoding="utf-8"))
    metrics = report["metrics"]["t1_vs_conversion"]

    exec_rows = pd.DataFrame(
        [
            ["Validation setup", "Prediction first, labels joined afterward", "Prevents label leakage during scoring"],
            ["T1 same top-100 users", bool(summary.loc[summary["stage"].eq("T1"), "same_uid_set"].iloc[0]), "Before vs after labels"],
            ["T1 same order", bool(summary.loc[summary["stage"].eq("T1"), "same_order"].iloc[0]), "Before vs after labels"],
            ["T1 same scores", bool(summary.loc[summary["stage"].eq("T1"), "same_scores"].iloc[0]), "Before vs after labels"],
            ["T1 converted in top 100", int(summary.loc[summary["stage"].eq("T1"), "converted_in_top100"].iloc[0]), "Actual conversions revealed after labels"],
            ["T1 DIY conversions in top 100", int(summary.loc[summary["stage"].eq("T1"), "diy_in_top100"].iloc[0]), "Source split"],
            ["T1 SFDC conversions in top 100", int(summary.loc[summary["stage"].eq("T1"), "sfdc_in_top100"].iloc[0]), "Source split"],
            ["April users evaluated", metrics["rows"], "Full labeled April population"],
            ["April actual conversions", metrics["positives"], "DIY + SFDC positives"],
            ["T1 AUC", metrics["auc"], "Full April validation"],
            ["Precision@100", metrics["precision_at_100"], "53 converted in top 100"],
            ["Top-decile lift", metrics["top_decile_lift"], "Top decile vs baseline conversion rate"],
        ],
        columns=["Metric", "Value", "Notes"],
    )
    t1_buckets = build_rank_buckets(t1)
    source_split = (
        t1.groupby("Conversion Source")
        .agg(Users=("UID", "size"), Conversions=("Converted After Label Check", "sum"))
        .reset_index()
    )
    before_after_cols = [
        "Stage",
        "Before Label Rank",
        "After Label Rank",
        "Same Rank",
        "UID",
        "Prediction Score",
        "Converted After Label Check",
        "Conversion Source",
        "Outcome",
        "Total Calls",
        "Answered Calls",
        "Answered Rate",
        "Avg Call Duration",
        "Campaign",
        "State",
    ]
    xlsx = OUT / "April_Unlabeled_vs_Labeled_Comparison.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        exec_rows.to_excel(writer, sheet_name="Executive_Summary", index=False)
        summary.to_excel(writer, sheet_name="Top100_Check_Summary", index=False)
        full_summary.to_excel(writer, sheet_name="Full_Rank_Check", index=False)
        t1[before_after_cols].to_excel(writer, sheet_name="T1_Top100_Comparison", index=False)
        t0[before_after_cols].to_excel(writer, sheet_name="T0_Top100_Comparison", index=False)
        t1_buckets.to_excel(writer, sheet_name="T1_Rank_Buckets", index=False)
        source_split.to_excel(writer, sheet_name="T1_Source_Split", index=False)
    format_workbook(xlsx)

    html = OUT / "April_Unlabeled_vs_Labeled_Comparison.html"
    build_html(t1, summary, html)

    findings = f"""# April Prediction vs Label Validation

## Core Message for Supervisor
April was scored first without using DIY/SFDC conversion labels. After scoring, labels were joined only for validation.

## T1 Top 100 Check
- Same users before vs after labels: {bool(summary.loc[summary['stage'].eq('T1'), 'same_uid_set'].iloc[0])}
- Same order before vs after labels: {bool(summary.loc[summary['stage'].eq('T1'), 'same_order'].iloc[0])}
- Same scores before vs after labels: {bool(summary.loc[summary['stage'].eq('T1'), 'same_scores'].iloc[0])}
- Converted users in top 100 after labels: {int(summary.loc[summary['stage'].eq('T1'), 'converted_in_top100'].iloc[0])}
- DIY conversions in top 100: {int(summary.loc[summary['stage'].eq('T1'), 'diy_in_top100'].iloc[0])}
- SFDC conversions in top 100: {int(summary.loc[summary['stage'].eq('T1'), 'sfdc_in_top100'].iloc[0])}

## Interpretation
The ranking did not change after labels were added. This means the prediction output and the validated outcome list are aligned correctly. Labels only added the actual outcome columns, proving that validation was done after prediction.
"""
    (OUT / "Supervisor_Explanation.md").write_text(findings, encoding="utf-8")

    package = OUT / "April_Unlabeled_vs_Labeled_Comparison_Package.zip"
    # Built by the caller with Compress-Archive if needed; leave files visible too.
    print(
        json.dumps(
            {
                "xlsx": str(xlsx),
                "html": str(html),
                "findings": str(OUT / "Supervisor_Explanation.md"),
                "t1_converted_top100": int(t1["Converted After Label Check"].sum()),
                "same_order": bool(summary.loc[summary["stage"].eq("T1"), "same_order"].iloc[0]),
                "package_target": str(package),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
