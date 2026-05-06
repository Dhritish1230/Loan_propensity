import hashlib
import json
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(os.getenv("LOAN_PROPENSITY_WORKSPACE", Path(__file__).resolve().parents[2])).resolve()
SOURCE = ROOT / "multi_month_training" / "test_outputs" / "april_t1_evaluated_predictions.csv"
OUT = ROOT / "per_call_behavior"
LABEL = "converted_full"
SCORE = "t1_loan_conversion_score"


def anon_id(value: str) -> str:
    return "USER_" + hashlib.sha256(str(value).encode("utf-8")).hexdigest()[:10].upper()


def pct(value: float) -> str:
    return f"{value * 100:.2f}%"


def call_bucket(value: float) -> str:
    v = int(value or 0)
    if v <= 7:
        return str(v)
    if v <= 10:
        return "8-10"
    if v <= 15:
        return "11-15"
    if v <= 20:
        return "16-20"
    return "21+"


def answered_bucket(value: float) -> str:
    v = int(value or 0)
    if v <= 5:
        return str(v)
    if v <= 8:
        return "6-8"
    return "9+"


def duration_bucket(row: pd.Series) -> str:
    calls = float(row.get("total_calls", 0) or 0)
    answered = float(row.get("answered_calls", 0) or 0)
    duration = float(row.get("avg_call_duration", 0) or 0)
    if calls <= 0:
        return "No calls"
    if answered <= 0 or duration <= 0:
        return "No answer / 0s"
    if duration < 10:
        return "1-9s"
    if duration < 30:
        return "10-29s"
    if duration < 60:
        return "30-59s"
    if duration < 120:
        return "60-119s"
    return "120s+"


def answered_rate_bucket(value: float) -> str:
    v = float(value or 0)
    if v <= 0:
        return "0%"
    if v < 0.25:
        return "1-24%"
    if v < 0.50:
        return "25-49%"
    if v < 0.75:
        return "50-74%"
    if v < 1.0:
        return "75-99%"
    return "100%"


def summarize(df: pd.DataFrame, group_col: str, order: list[str] | None = None) -> pd.DataFrame:
    baseline = df[LABEL].mean()
    out = (
        df.groupby(group_col, dropna=False)
        .agg(
            Users=("uid", "size"),
            Conversions=(LABEL, "sum"),
            Conversion_Rate=(LABEL, "mean"),
            Expected_Conversions=(SCORE, "sum"),
            Avg_T1_Score=(SCORE, "mean"),
            Avg_T0_Score=("t0_call_targeting_score", "mean"),
            Total_Calls_Made=("total_calls", "sum"),
            Avg_Total_Calls=("total_calls", "mean"),
            Avg_Answered_Calls=("answered_calls", "mean"),
            Avg_Answered_Rate=("answered_rate", "mean"),
            Avg_Call_Duration=("avg_call_duration", "mean"),
        )
        .reset_index()
        .rename(columns={group_col: "Bucket"})
    )
    out["Lift_vs_Baseline"] = out["Conversion_Rate"] / baseline
    out["Conversions_per_1000_Calls"] = out.apply(
        lambda r: (r["Conversions"] / r["Total_Calls_Made"] * 1000) if r["Total_Calls_Made"] else 0,
        axis=1,
    )
    if order:
        out["__order"] = out["Bucket"].map({label: idx for idx, label in enumerate(order)}).fillna(999)
        out = out.sort_values("__order").drop(columns="__order")
    return out


def minimum_depth_curve(df: pd.DataFrame, col: str, max_depth: int, label: str) -> pd.DataFrame:
    rows = []
    baseline = df[LABEL].mean()
    for depth in range(0, max_depth + 1):
        sub = df.loc[pd.to_numeric(df[col], errors="coerce").fillna(0) >= depth]
        rows.append(
            {
                "Depth_Type": label,
                "Minimum_Depth": depth,
                "Users_Reached": len(sub),
                "Conversions": int(sub[LABEL].sum()),
                "Conversion_Rate": float(sub[LABEL].mean()) if len(sub) else 0,
                "Lift_vs_Baseline": (float(sub[LABEL].mean()) / baseline) if len(sub) and baseline else 0,
                "Avg_T1_Score": float(sub[SCORE].mean()) if len(sub) else 0,
                "Avg_Call_Duration": float(sub["avg_call_duration"].mean()) if len(sub) else 0,
            }
        )
    return pd.DataFrame(rows)


def exact_marginal_table(df: pd.DataFrame) -> pd.DataFrame:
    baseline = df[LABEL].mean()
    exact = (
        df.assign(Exact_Total_Calls=pd.to_numeric(df["total_calls"], errors="coerce").fillna(0).astype(int).clip(upper=25))
        .groupby("Exact_Total_Calls")
        .agg(
            Users=("uid", "size"),
            Conversions=(LABEL, "sum"),
            Conversion_Rate=(LABEL, "mean"),
            Avg_T1_Score=(SCORE, "mean"),
            Avg_Answered_Calls=("answered_calls", "mean"),
            Avg_Call_Duration=("avg_call_duration", "mean"),
        )
        .reset_index()
    )
    exact["Lift_vs_Baseline"] = exact["Conversion_Rate"] / baseline
    exact["Conversion_Rate_Change_vs_Previous"] = exact["Conversion_Rate"].diff()
    exact["Exact_Total_Calls"] = exact["Exact_Total_Calls"].astype(str)
    exact.loc[exact["Exact_Total_Calls"].eq(25), "Exact_Total_Calls"] = "25+"
    return exact


def add_summary_sheet(writer, df: pd.DataFrame, tables: dict[str, pd.DataFrame]) -> None:
    baseline = df[LABEL].mean()
    total_calls = df["total_calls"].sum()
    answered_calls = df["answered_calls"].sum()
    best_call_bucket = tables["Total_Call_Bucket"].loc[
        tables["Total_Call_Bucket"].query("Users >= 500")["Conversion_Rate"].idxmax()
    ]
    best_duration = tables["Duration_Bucket"].loc[
        tables["Duration_Bucket"].query("Users >= 500")["Conversion_Rate"].idxmax()
    ]
    kpis = pd.DataFrame(
        [
            ["Users analyzed", len(df), "April labeled holdout users"],
            ["Actual conversions", int(df[LABEL].sum()), "DIY + SFDC positives"],
            ["Baseline conversion", baseline, "All users"],
            ["Total calls made", int(total_calls), "Joined call records"],
            ["Answered calls", int(answered_calls), "Answered call count"],
            ["Answered-call share", answered_calls / total_calls if total_calls else 0, "Answered calls / total calls"],
            ["Best call-count bucket", best_call_bucket["Bucket"], f"{pct(best_call_bucket['Conversion_Rate'])} conversion"],
            ["Best duration bucket", best_duration["Bucket"], f"{pct(best_duration['Conversion_Rate'])} conversion"],
        ],
        columns=["Metric", "Value", "Notes"],
    )
    kpis.to_excel(writer, sheet_name="Summary", index=False)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="102027")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="EAF7F2")
    thin = Side(style="thin", color="D7DEE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col), 38)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(row=1, column=cell.column).value
                if header and any(token in str(header) for token in ["Rate", "Lift", "Precision", "Recall", "Coverage"]):
                    cell.number_format = "0.00%"
                if header and any(token in str(header) for token in ["Score", "Duration"]):
                    cell.number_format = "0.000"

    summary = wb["Summary"]
    summary.insert_rows(1, 2)
    summary["A1"] = "Per-Call Behavior Analysis"
    summary["A1"].font = Font(size=20, bold=True, color="102027")
    summary["A2"] = "Observational view of how conversion changes by call count, answered calls, and call duration."
    summary["A2"].font = Font(italic=True, color="60717C")
    for col in range(1, 4):
        summary.cell(row=3, column=col).fill = header_fill
        summary.cell(row=3, column=col).font = header_font
    summary["A1"].fill = title_fill

    def add_bar(sheet_name: str, title: str, data_col: int, anchor: str) -> None:
        ws = wb[sheet_name]
        chart = BarChart()
        chart.title = title
        chart.y_axis.title = "Conversion Rate"
        chart.x_axis.title = "Bucket"
        data = Reference(ws, min_col=data_col, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, anchor)

    add_bar("Total_Call_Bucket", "Conversion Rate by Total Calls", 4, "N2")
    add_bar("Answered_Call_Bucket", "Conversion Rate by Answered Calls", 4, "N2")
    add_bar("Duration_Bucket", "Conversion Rate by Avg Call Duration", 4, "N2")

    ws = wb["Min_Call_Depth"]
    chart = LineChart()
    chart.title = "Conversion Rate When Users Receive At Least N Calls"
    chart.y_axis.title = "Conversion Rate"
    chart.x_axis.title = "Minimum Calls"
    data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "J2")

    for sheet in ["Total_Call_Bucket", "Answered_Call_Bucket", "Duration_Bucket", "Answered_Rate_Bucket"]:
        ws = wb[sheet]
        for col_name in ["Conversion_Rate", "Lift_vs_Baseline", "Conversions_per_1000_Calls"]:
            headers = [cell.value for cell in ws[1]]
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                ws.conditional_formatting.add(
                    f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{ws.max_row}",
                    ColorScaleRule(start_type="min", start_color="FEE4E2", mid_type="percentile", mid_value=50, mid_color="FFF3CD", end_type="max", end_color="DFF7E8"),
                )

    wb.save(path)


def build_html(df: pd.DataFrame, tables: dict[str, pd.DataFrame], path: Path) -> None:
    baseline = df[LABEL].mean()
    total_calls = int(df["total_calls"].sum())
    answered_calls = int(df["answered_calls"].sum())
    total_table = tables["Total_Call_Bucket"]
    duration_table = tables["Duration_Bucket"]
    min_depth = tables["Min_Call_Depth"]

    def table_html(frame: pd.DataFrame, cols: list[str]) -> str:
        rows = []
        for _, row in frame[cols].iterrows():
            cells = []
            for col in cols:
                val = row[col]
                if isinstance(val, float) and ("Rate" in col or "Lift" in col):
                    val = pct(val)
                elif isinstance(val, float):
                    val = f"{val:,.3f}"
                elif isinstance(val, int):
                    val = f"{val:,}"
                cells.append(f"<td>{val}</td>")
            rows.append("<tr>" + "".join(cells) + "</tr>")
        return "<table><thead><tr>" + "".join(f"<th>{c}</th>" for c in cols) + "</tr></thead><tbody>" + "".join(rows) + "</tbody></table>"

    def bars(frame: pd.DataFrame, label_col: str = "Bucket") -> str:
        max_rate = frame["Conversion_Rate"].max() or 1
        html = []
        for _, row in frame.iterrows():
            width = max(3, row["Conversion_Rate"] / max_rate * 100)
            html.append(
                f"""<div class="bar-row"><b>{row[label_col]}</b><div class="track"><span style="width:{width:.1f}%"></span></div><strong>{pct(row['Conversion_Rate'])}</strong><em>{int(row['Users']):,} users</em></div>"""
            )
        return "\n".join(html)

    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Per-Call Behavior Analysis</title>
<style>
body{{margin:0;font-family:Bahnschrift,Candara,'Trebuchet MS',sans-serif;color:#102027;background:linear-gradient(135deg,#fff9ec,#eaf7f2)}}
.wrap{{width:min(1320px,calc(100% - 34px));margin:0 auto;padding:32px 0 60px}}
.hero,.panel,.card{{background:rgba(255,255,255,.82);border:1px solid rgba(16,32,39,.12);border-radius:28px;box-shadow:0 24px 70px rgba(23,50,76,.13);padding:24px;margin-top:18px}}
h1,h2{{font-family:Georgia,'Palatino Linotype',serif;letter-spacing:-.035em}}h1{{font-size:64px;line-height:.95;margin:0}}p{{color:#60717c;line-height:1.55}}
.grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px}}.card strong{{display:block;font-size:38px;margin-top:8px}}.card span{{color:#60717c}}
.two{{display:grid;grid-template-columns:1fr 1fr;gap:18px}}.bar-row{{display:grid;grid-template-columns:100px 1fr 90px 110px;gap:10px;align-items:center;margin:11px 0}}.track{{height:16px;border-radius:999px;background:#edf1f0;overflow:hidden}}.track span{{display:block;height:100%;border-radius:999px;background:linear-gradient(90deg,#0f766e,#d9911b)}}
table{{width:100%;border-collapse:collapse;background:white;border-radius:18px;overflow:hidden}}th,td{{padding:10px;border-bottom:1px solid #edf1f0;text-align:left}}th{{background:#102027;color:white;font-size:12px;text-transform:uppercase;letter-spacing:.08em}}
.note{{background:#fff3cd;border-color:#ffe69c}}@media(max-width:900px){{.grid,.two{{grid-template-columns:1fr}}h1{{font-size:42px}}}}
</style></head><body><main class="wrap">
<section class="hero"><p style="font-weight:900;color:#0f766e;text-transform:uppercase;letter-spacing:.16em">Mentor analysis</p><h1>Per-call behavior and conversion movement</h1><p>This report shows how conversion changes with total calls, answered calls, answered rate, and average call duration. It is observational, so it helps guide calling strategy but does not prove that extra calls alone cause conversions.</p></section>
<section class="grid">
<article class="card"><span>Users analyzed</span><strong>{len(df):,}</strong></article>
<article class="card"><span>Actual conversions</span><strong>{int(df[LABEL].sum()):,}</strong></article>
<article class="card"><span>Baseline conversion</span><strong>{pct(baseline)}</strong></article>
<article class="card"><span>Answered call share</span><strong>{pct(answered_calls / total_calls if total_calls else 0)}</strong></article>
</section>
<section class="panel note"><h2>How to interpret this</h2><p>If conversion rises with more calls or longer answered calls, that suggests call engagement is a strong signal. But users receiving more calls may already be different from users receiving fewer calls, so this should be used as business evidence, not causal proof. A randomized call-depth experiment would be needed to prove incremental call impact.</p></section>
<section class="panel two"><div><h2>Conversion by total calls</h2>{bars(total_table)}</div><div><h2>Conversion by avg duration</h2>{bars(duration_table)}</div></section>
<section class="panel"><h2>Minimum call-depth curve</h2>{table_html(min_depth, ['Minimum_Depth','Users_Reached','Conversions','Conversion_Rate','Lift_vs_Baseline','Avg_T1_Score'])}</section>
<section class="panel"><h2>Total call bucket table</h2>{table_html(total_table, ['Bucket','Users','Conversions','Conversion_Rate','Lift_vs_Baseline','Conversions_per_1000_Calls','Avg_T1_Score','Avg_Call_Duration'])}</section>
</main></body></html>"""
    path.write_text(html, encoding="utf-8")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    df = pd.read_csv(SOURCE, low_memory=False)
    for col in ["total_calls", "answered_calls", "answered_rate", "avg_call_duration", LABEL, SCORE, "t0_call_targeting_score"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["Anon_User_ID"] = df["uid"].map(anon_id)
    df["Total_Call_Bucket"] = df["total_calls"].map(call_bucket)
    df["Answered_Call_Bucket"] = df["answered_calls"].map(answered_bucket)
    df["Duration_Bucket"] = df.apply(duration_bucket, axis=1)
    df["Answered_Rate_Bucket"] = df["answered_rate"].map(answered_rate_bucket)

    orders = {
        "Total_Call_Bucket": ["0", "1", "2", "3", "4", "5", "6", "7", "8-10", "11-15", "16-20", "21+"],
        "Answered_Call_Bucket": ["0", "1", "2", "3", "4", "5", "6-8", "9+"],
        "Duration_Bucket": ["No calls", "No answer / 0s", "1-9s", "10-29s", "30-59s", "60-119s", "120s+"],
        "Answered_Rate_Bucket": ["0%", "1-24%", "25-49%", "50-74%", "75-99%", "100%"],
    }

    tables = {
        "Total_Call_Bucket": summarize(df, "Total_Call_Bucket", orders["Total_Call_Bucket"]),
        "Answered_Call_Bucket": summarize(df, "Answered_Call_Bucket", orders["Answered_Call_Bucket"]),
        "Duration_Bucket": summarize(df, "Duration_Bucket", orders["Duration_Bucket"]),
        "Answered_Rate_Bucket": summarize(df, "Answered_Rate_Bucket", orders["Answered_Rate_Bucket"]),
        "Min_Call_Depth": minimum_depth_curve(df, "total_calls", 12, "Total Calls"),
        "Min_Answered_Depth": minimum_depth_curve(df, "answered_calls", 8, "Answered Calls"),
        "Exact_Marginal_Calls": exact_marginal_table(df),
    }

    matrix = (
        df.groupby(["Total_Call_Bucket", "Duration_Bucket"])
        .agg(Users=("uid", "size"), Conversions=(LABEL, "sum"), Conversion_Rate=(LABEL, "mean"))
        .reset_index()
    )
    tables["Call_Count_x_Duration"] = matrix

    top_users = df.sort_values(SCORE, ascending=False).head(250)[
        [
            "Anon_User_ID",
            "state",
            "campaign_id",
            "total_calls",
            "answered_calls",
            "answered_rate",
            "avg_call_duration",
            SCORE,
            LABEL,
            "Total_Call_Bucket",
            "Duration_Bucket",
        ]
    ].rename(
        columns={
            "state": "State",
            "campaign_id": "Campaign",
            "total_calls": "Total_Calls",
            "answered_calls": "Answered_Calls",
            "answered_rate": "Answered_Rate",
            "avg_call_duration": "Avg_Call_Duration",
            SCORE: "T1_Score",
            LABEL: "Converted",
        }
    )
    tables["Top_250_Users_Anon"] = top_users

    for name, table in tables.items():
        table.to_csv(OUT / f"{name}.csv", index=False)

    xlsx_path = OUT / "Per_Call_Behavior_Report.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        add_summary_sheet(writer, df, tables)
        for name, table in tables.items():
            table.to_excel(writer, sheet_name=name[:31], index=False)
    format_workbook(xlsx_path)

    html_path = OUT / "Per_Call_Behavior_Dashboard.html"
    build_html(df, tables, html_path)

    baseline = df[LABEL].mean()
    best_call = tables["Total_Call_Bucket"].query("Users >= 500").sort_values("Conversion_Rate", ascending=False).iloc[0]
    best_duration = tables["Duration_Bucket"].query("Users >= 500").sort_values("Conversion_Rate", ascending=False).iloc[0]
    findings = f"""# Per-Call Behavior Findings

## What We Analyzed
April holdout users were grouped by:
- Total number of calls received.
- Number of answered calls.
- Answered-call rate.
- Average call duration.
- Minimum call depth, e.g. users who received at least 1, 2, 3, ... calls.

## Key Numbers
- Users analyzed: {len(df):,}
- Actual conversions: {int(df[LABEL].sum()):,}
- Baseline conversion rate: {pct(baseline)}
- Total calls observed: {int(df['total_calls'].sum()):,}
- Answered calls observed: {int(df['answered_calls'].sum()):,}

## Best Observed Buckets
- Best total-call bucket with at least 500 users: `{best_call['Bucket']}` calls, conversion rate {pct(best_call['Conversion_Rate'])}, lift {best_call['Lift_vs_Baseline']:.2f}x.
- Best average-duration bucket with at least 500 users: `{best_duration['Bucket']}`, conversion rate {pct(best_duration['Conversion_Rate'])}, lift {best_duration['Lift_vs_Baseline']:.2f}x.

## Important Caveat
This analysis is observational. It shows that call behavior is associated with conversion, but it does not prove that simply increasing calls will cause the same conversion lift. Users who receive more calls may differ from users who receive fewer calls. To prove causal impact, run a controlled call-depth experiment.

## Mentor-Friendly Recommendation
Use this analysis to decide a practical call strategy:
1. Prioritize T0 users for first outreach.
2. Track whether calls are answered and whether duration crosses useful engagement thresholds.
3. Use T1 after calls to prioritize follow-up.
4. Monitor whether additional calls after a certain depth are still producing enough incremental conversions per 1,000 calls.
"""
    (OUT / "Per_Call_Behavior_Findings.md").write_text(findings, encoding="utf-8")

    print(
        json.dumps(
            {
                "output_dir": str(OUT),
                "xlsx": str(xlsx_path),
                "html": str(html_path),
                "users": len(df),
                "baseline_conversion": baseline,
                "best_call_bucket": str(best_call["Bucket"]),
                "best_duration_bucket": str(best_duration["Bucket"]),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
